import time
import random
from decimal import Decimal
import openpyxl
from web3 import Web3

# Импортируем все настройки из config.py
import config


def main():
    web3 = Web3(Web3.HTTPProvider(config.RPC_URL))
    if not web3.is_connected():
        raise ConnectionError("Не удалось подключиться к RPC. Проверьте RPC_URL в config.py.")

    # Открываем Excel
    wb = openpyxl.load_workbook(config.EXCEL_FILE)
    sheet = wb[config.SHEET_NAME]

    # Создаём объект контракта
    token_contract = web3.eth.contract(
        address=Web3.to_checksum_address(config.TOKEN_CONTRACT_ADDRESS),
        abi=config.ERC20_ABI
    )

    # Узнаём decimals у токена
    try:
        token_decimals = token_contract.functions.decimals().call()
    except Exception as e:
        print(f"Не удалось получить decimals у контракта: {e}")
        print("Допустим decimals = 18.")
        token_decimals = 18

    # Фактор для пересчёта из «человеческого» формата в wei
    decimals_factor = Decimal(10) ** Decimal(token_decimals)

    # Перевод Gwei -> wei (для EIP-1559)
    max_fee_per_gas_wei = int(config.MAX_FEE_PER_GAS_GWEI * (10**9))
    max_priority_fee_per_gas_wei = int(config.MAX_PRIORITY_FEE_PER_GAS_GWEI * (10**9))

    # min_amount_transfer в wei
    min_transfer_wei = int(Decimal(config.min_amount_transfer) * decimals_factor)

    # Проходим по всем строкам Excel
    for row in range(1, sheet.max_row + 1):
        private_key = sheet.cell(row=row, column=1).value
        to_address = sheet.cell(row=row, column=2).value

        # Пропускаем пустые ячейки
        if not private_key or not to_address:
            continue

        try:
            # Формируем аккаунт
            account = web3.eth.account.from_key(private_key)
            sender_address = account.address

            # Узнаём баланс токенов
            token_balance = token_contract.functions.balanceOf(sender_address).call()

            # Определяем сумму к отправке
            if config.transfer_all_balance:
                # Отправляем весь баланс, кроме случайного «хвоста»
                keep_tokens = random.uniform(config.keep_value_from, config.keep_value_to)
                keep_wei = int(Decimal(str(keep_tokens)) * decimals_factor)
                amount_to_send = token_balance - keep_wei
                if amount_to_send < 0:
                    amount_to_send = 0
            else:
                # Случайная сумма из [amount_from, amount_to]
                rand_amount = random.uniform(config.amount_from, config.amount_to)
                amount_to_send = int(Decimal(str(rand_amount)) * decimals_factor)

            # Проверка: если сумма меньше min_transfer_wei — пропускаем
            if amount_to_send < min_transfer_wei:
                print(f"[{sender_address}] Сумма {amount_to_send} (wei) < min_amount_transfer, пропускаем.")
                continue

            # Проверка, хватает ли токенов
            if token_balance < amount_to_send:
                print(f"[{sender_address}] Недостаточно токенов. Баланс={token_balance}, нужно={amount_to_send}")
                continue

            # Проверка, сумма не 0
            if amount_to_send <= 0:
                print(f"[{sender_address}] Сумма для отправки = 0, пропускаем.")
                continue

            # Nonce
            nonce = web3.eth.get_transaction_count(sender_address)

            # Оценка газа (Web3.py 6+ → estimate_gas())
            try:
                estimated_gas = token_contract.functions.transfer(
                    Web3.to_checksum_address(to_address),
                    amount_to_send
                ).estimate_gas({"from": sender_address})
            except Exception as e_gas:
                print(f"[{sender_address}] Ошибка estimate_gas: {e_gas}")
                continue
            
            latest_block = web3.eth.get_block('latest')
            base_fee = latest_block.baseFeePerGas  # в wei
            priority_fee = max_priority_fee_per_gas_wei  # то, что вы задали
            approx_fee_per_gas = base_fee + priority_fee

            # Добавим небольшой запас в 20% (safety margin)
            approx_fee_per_gas = int(approx_fee_per_gas * 1.2)
            # Проверяем нативный баланс для оплаты газа
            native_balance = web3.eth.get_balance(sender_address)
            # Грубая верхняя оценка при EIP-1559
            estimated_fee = estimated_gas * approx_fee_per_gas
            if native_balance < estimated_fee:
                print(f"[{sender_address}] Недостаточно нативного баланса для газа. Нужно ~{estimated_fee}, есть {native_balance}")
                continue

            # Формируем транзакцию (EIP-1559).
            # Если сеть не поддерживает EIP-1559, замените на 'gasPrice': ...
            tx = token_contract.functions.transfer(
                Web3.to_checksum_address(to_address),
                amount_to_send
            ).build_transaction({
                "chainId": web3.eth.chain_id,
                "nonce": nonce,
                "gas": estimated_gas,
                "maxFeePerGas": approx_fee_per_gas,
                "maxPriorityFeePerGas": priority_fee,
            })

            signed_tx = web3.eth.account.sign_transaction(tx, private_key=private_key)

            # Отправляем (Web3.py 6+ → send_raw_transaction + signed_tx.raw_transaction)
            tx_hash = web3.eth.send_raw_transaction(signed_tx.raw_transaction)

            # Для лога переводим сумму в «человеческий» формат
            human_amount = Decimal(amount_to_send) / decimals_factor
            print(
                f"[{sender_address}] --> [{to_address}] "
                f"Отправляем {human_amount} токенов. Tx hash: {tx_hash.hex()}"
            )

            # Задержка
            time.sleep(config.DELAY_BETWEEN_TXS)

        except Exception as e:
            print(f"Ошибка при отправке с [{sender_address}] на [{to_address}]: {e}")


if __name__ == "__main__":
    main()
